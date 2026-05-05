"""Generate relationship.xlsx mirroring the screenshot of relationship.xls."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUTPUT = Path(__file__).parent / "relationship.xlsx"

HEADERS = [
    "Entity",
    "Model View",
    "Base Bde trigger",
    "Base BDE Time Travel",
    "Base BDE Dependancy",
    "Dependent BDE Time Travel",
    "Parallel lookup",
    "Reference lookup",
    "Base Bde Time travel status",
]

# (entity, model_view, base_bde_trigger, base_bde_time_travel,
#  base_bde_dependancy, dep_bde_time_travel, parallel_lookup,
#  reference_lookup, status)
TRANSACTION_ROWS = [
    ("Transaction-FXTR",    "bde_books", "1 Day", "doc_fxtr",                "1 Day", "Asset,Valuation,Asset_class,Account"),
    ("Transaction-INPAY",   "bde_books", "1 Day", "doc_inpay",               "1 Day", "Asset_class,Account"),
    ("Transaction-INTR",    "bde_books", "1 Day", "doc_intr",                "1 Day", "Asset_class,Account"),
    ("Transaction-LIMIT",   "bde_books", "1 Day", "doc_limit",               "1 Day", "Asset_class,Account"),
    ("Transaction-LOAN",    "bde_books", "1 Day", "doc_loan",                "1 Day", "Asset_class,Account"),
    ("Transaction-MMKT",    "bde_books", "1 Day", "doc_mmkt",                "1 Day", "Asset_class,Account"),
    ("Transaction-PAY",     "bde_books", "1 Day", "doc_pay",                 "1 Day", "Order,Asset,Asset_class,Account"),
    ("Transaction-SECTRX2", "bde_books", "1 Day", "doc_sectrx2,doc_secevt2", "1 Day", "Asset,Asset_class,Account"),
    ("Transaction-SETTLE",  "bde_books", "1 Day", "doc_settle",              "1 Day", "Asset_class,Account"),
    ("Transaction-STEX",    "bde_books", "1 Day", "doc_stex",                "1 Day", "Account,Valuation,Asset,Asset_class,Account"),
    ("Transaction-XFER",    "bde_books", "1 Day", "doc_xfer",                "1 Day", "Asset_class,Account"),
    ("Transaction-XFERFEE", "bde_books", "1 Day", "doc_xferfee",             "1 Day", "Asset,Order,Asset_class,Account"),
    ("Transaction-XFERMON", "bde_books", "1 day", "doc_xfermon",             "1 day", "Asset,Asset_class,Account"),
]

CHARGE_ROWS = [
    ("Transaction-Charge-FXTR",    "doc_fxtr",    "1 Day"),
    ("Transaction-Charge-INPAY",   "doc_inpay",   "1 Day"),
    ("Transaction-Charge-INTR",    "doc_intr",    "1 Day"),
    ("Transaction-Charge-LIMIT",   "doc_limit",   "1 Day"),
    ("Transaction-Charge-LOAN",    "doc_loan",    "1 Day"),
    ("Transaction-Charge-MMKT",    "doc_mmkt",    "1 Day"),
    ("Transaction-Charge-PAY",     "doc_pay",     "1 Day"),
    ("Transaction-Charge-SECTRX2", "doc_sectrx2", "1 Day"),
    ("Transaction-Charge-SETTLE",  "doc_settle",  "1 Day"),
    ("Transaction-Charge-STEX",    "doc_stex",    "1 Day"),
    ("Transaction-Charge-XFER",    "doc_xfer",    "1 Day"),
    ("Transaction-Charge-XFERMON", "doc_xfermon", "1 Day"),
]

ORDER_ROWS = [
    ("Order-FXSW",       "doc_fxsw",                  "1 Day", "doc_fxtr,bde_books",      "1 Day", "Asset,Account"),
    ("Order-FXTR",       "doc_fxtr",                  "1 day", "bde_books",               "1 day", "Asset,Account"),
    ("Order-MARKETFILL", "doc_form___bw_market_fill", "1 day", "doc_stex,bde_books",      "1 day", "Asset,Account"),
    ("Order-PAY",        "doc_pay",                   "1 day", "bde_books",               "1 day", "Asset,Account"),
    ("Order-SECEVT2",    "doc_secevt2",               "1 day", "",                        "",      "Asset"),
    ("Order-SECTRX2",    "doc_sectrx2",               "1 day", "bde_books,doc_secevt2",   "1 day", "Asset,Account"),
    ("Order-SETTLE",     "doc_settle",                "1 day", "bde_books",               "1 day", "Asset,Account"),
    ("Order-STEX",       "doc_stex",                  "1 day", "bde_books",               "1 day", "Asset,Account"),
    ("Order-XFER",       "doc_xfer",                  "1 day", "bde_books",               "1 day", "Asset,Account"),
    ("Order-XFERMON",    "doc_xfermon",               "1 day", "bde_books",               "1 day", "Asset,Account"),
]


def build_full_rows():
    rows = []
    for r in TRANSACTION_ROWS:
        rows.append(("Transaction", *r, "", "Completed"))
    for r in CHARGE_ROWS:
        # charge rows: dependancy empty, dep_time "N/A", parallel "Asset", reference "avq_codes_tab"
        model_view, trigger, time_travel = r
        rows.append((
            "Transaction-Charge", model_view, trigger, time_travel,
            "", "N/A", "Asset", "avq_codes_tab", "Completed",
        ))
    for r in ORDER_ROWS:
        rows.append(("Order", *r, "", "Completed"))
    return rows


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="4F4F6F")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_xlsx() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relationships"

    thin = Side(border_style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header(cell)
        cell.border = border

    rows = build_full_rows()
    # Determine merge ranges for column A (Entity) — merge consecutive same-entity rows
    current_entity = None
    merge_start = 2
    excel_row = 2

    for row_data in rows:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        entity = row_data[0]
        if entity != current_entity:
            if current_entity is not None and excel_row - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=excel_row - 1, end_column=1)
                ws.cell(row=merge_start, column=1).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws.merge_cells(start_row=merge_start, start_column=9, end_row=excel_row - 1, end_column=9)
                ws.cell(row=merge_start, column=9).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            current_entity = entity
            merge_start = excel_row
        excel_row += 1

    # final merge for the last entity block
    if excel_row - 1 > merge_start:
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=excel_row - 1, end_column=1)
        ws.cell(row=merge_start, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.merge_cells(start_row=merge_start, start_column=9, end_row=excel_row - 1, end_column=9)
        ws.cell(row=merge_start, column=9).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    widths = [22, 32, 28, 22, 28, 22, 42, 18, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"

    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = write_xlsx()
    print(f"Wrote {path}")
