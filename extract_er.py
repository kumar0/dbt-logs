#!/usr/bin/env python3
"""
Extract entities and relationships from dbt artifacts.

Run this from the directory containing catalog.json and manifest.json:
    python extract_er.py

Inputs (in current dir):
    catalog.json   - entities (models/sources) and their columns
    manifest.json  - relationships (refs/sources + FK tests). Optional.

Outputs (in current dir):
    er_model.json        - full structured dump
    entities.csv         - one row per entity
    columns.csv          - one row per column
    relationships.csv    - one row per edge
    er_model.xlsx        - same three tables as sheets
"""

import csv
import json
from pathlib import Path
from collections import defaultdict

CATALOG_PATH = "catalog.json"
MANIFEST_PATH = "manifest.json"
OUTPUT_JSON = "er_model.json"
OUTPUT_XLSX = "er_model.xlsx"
ENTITIES_CSV = "entities.csv"
COLUMNS_CSV = "columns.csv"
RELATIONSHIPS_CSV = "relationships.csv"


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_entities(catalog):
    """Pull entities (tables/views) and their columns from catalog.json."""
    entities = []

    # catalog.json has two top-level node groups: "nodes" (models, seeds, snapshots)
    # and "sources" (source tables).
    for section in ("nodes", "sources"):
        for unique_id, node in catalog.get(section, {}).items():
            meta = node.get("metadata", {})
            stats = node.get("stats", {})

            columns = []
            for col_name, col in node.get("columns", {}).items():
                columns.append({
                    "name": col.get("name", col_name),
                    "type": col.get("type"),
                    "index": col.get("index"),
                    "comment": col.get("comment"),
                })

            # Sort columns by their index when available
            columns.sort(key=lambda c: (c["index"] is None, c["index"]))

            row_count = None
            if "row_count" in stats and isinstance(stats["row_count"], dict):
                row_count = stats["row_count"].get("value")

            entities.append({
                "unique_id": unique_id,
                "kind": section[:-1],          # "node" or "source"
                "type": meta.get("type"),      # BASE TABLE / VIEW / etc.
                "database": meta.get("database"),
                "schema": meta.get("schema"),
                "name": meta.get("name"),
                "comment": meta.get("comment"),
                "owner": meta.get("owner"),
                "row_count": row_count,
                "columns": columns,
            })

    return entities


def extract_relationships(manifest):
    """Pull relationships from manifest.json.

    Two sources of edges:
      1. Lineage edges from `depends_on.nodes` (model -> upstream model/source).
      2. Foreign-key edges from `relationships` tests in dbt tests.
    """
    if not manifest:
        return []

    rels = []
    nodes = manifest.get("nodes", {})
    sources = manifest.get("sources", {})

    # 1) Lineage edges
    for unique_id, node in nodes.items():
        if node.get("resource_type") not in ("model", "snapshot", "seed"):
            continue
        for upstream in node.get("depends_on", {}).get("nodes", []):
            rels.append({
                "kind": "lineage",
                "from": upstream,        # parent (referenced)
                "to": unique_id,         # child (referrer)
            })

    # 2) Foreign-key style relationships from `relationships` tests.
    #    These tests look like: test_relationships_<col>__<ref_col>__ref_<model>_
    for unique_id, node in nodes.items():
        if node.get("resource_type") != "test":
            continue
        test_meta = node.get("test_metadata") or {}
        if test_meta.get("name") != "relationships":
            continue

        kwargs = test_meta.get("kwargs", {}) or {}
        # The test is attached to a model/column; figure out which.
        # `attached_node` is the dbt-canonical field; fall back to depends_on parents.
        from_node = node.get("attached_node")
        from_column = kwargs.get("column_name") or node.get("column_name")

        # Target model is given as a Jinja string like "ref('dim_customer')".
        # The compiled `refs` / `sources` list is the reliable place to look.
        refs = node.get("refs") or []
        srcs = node.get("sources") or []

        to_node = None
        if refs:
            ref = refs[0]
            ref_name = ref["name"] if isinstance(ref, dict) else ref[-1]
            # Look up the unique_id of that model
            for nid, n in nodes.items():
                if n.get("resource_type") == "model" and n.get("name") == ref_name:
                    to_node = nid
                    break
        elif srcs:
            src = srcs[0]
            # sources entries look like ["source_name", "table_name"]
            if isinstance(src, list) and len(src) >= 2:
                for sid, s in sources.items():
                    if s.get("source_name") == src[0] and s.get("name") == src[1]:
                        to_node = sid
                        break

        to_column = kwargs.get("field")

        if from_node and to_node:
            rels.append({
                "kind": "foreign_key",
                "from": from_node,
                "from_column": from_column,
                "to": to_node,
                "to_column": to_column,
            })

    return rels


def print_summary(entities, relationships):
    print(f"\n=== Entities: {len(entities)} ===\n")
    by_kind = defaultdict(int)
    for e in entities:
        by_kind[e["kind"]] += 1
    for k, v in by_kind.items():
        print(f"  {k}s: {v}")

    print(f"\n=== Relationships: {len(relationships)} ===\n")
    by_kind = defaultdict(int)
    for r in relationships:
        by_kind[r["kind"]] += 1
    for k, v in by_kind.items():
        print(f"  {k}: {v}")

    # Show a sample
    print("\n--- Sample entities (first 5) ---")
    for e in entities[:5]:
        fq = ".".join(filter(None, [e["database"], e["schema"], e["name"]]))
        print(f"  [{e['kind']}] {fq}  ({len(e['columns'])} cols)")

    if relationships:
        print("\n--- Sample relationships (first 10) ---")
        for r in relationships[:10]:
            if r["kind"] == "foreign_key":
                print(f"  FK  {r['from']}.{r.get('from_column')}  ->  "
                      f"{r['to']}.{r.get('to_column')}")
            else:
                print(f"  LIN {r['from']}  ->  {r['to']}")


def write_csvs(entities, relationships):
    # entities.csv (one row per entity)
    with open(ENTITIES_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["unique_id", "kind", "type", "database", "schema", "name",
                    "owner", "row_count", "column_count", "comment"])
        for e in entities:
            w.writerow([
                e["unique_id"], e["kind"], e["type"], e["database"], e["schema"],
                e["name"], e["owner"], e["row_count"], len(e["columns"]),
                e["comment"],
            ])

    # columns.csv (one row per column)
    with open(COLUMNS_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["entity_unique_id", "entity_name", "column_index",
                    "column_name", "column_type", "column_comment"])
        for e in entities:
            for c in e["columns"]:
                w.writerow([
                    e["unique_id"], e["name"], c["index"], c["name"],
                    c["type"], c["comment"],
                ])

    # relationships.csv (one row per edge)
    with open(RELATIONSHIPS_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["kind", "from", "from_column", "to", "to_column"])
        for r in relationships:
            w.writerow([
                r["kind"], r["from"], r.get("from_column", ""),
                r["to"], r.get("to_column", ""),
            ])


def write_excel(entities, relationships):
    """Write the same three tables as sheets in one xlsx."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("(skipping Excel — install openpyxl to enable: pip install openpyxl)")
        return

    wb = Workbook()

    ws = wb.active
    ws.title = "entities"
    headers = ["unique_id", "kind", "type", "database", "schema", "name",
               "owner", "row_count", "column_count", "comment"]
    ws.append(headers)
    for e in entities:
        ws.append([
            e["unique_id"], e["kind"], e["type"], e["database"], e["schema"],
            e["name"], str(e["owner"]) if e["owner"] else "",
            e["row_count"], len(e["columns"]), e["comment"],
        ])

    ws2 = wb.create_sheet("columns")
    ws2.append(["entity_unique_id", "entity_name", "column_index",
                "column_name", "column_type", "column_comment"])
    for e in entities:
        for c in e["columns"]:
            ws2.append([
                e["unique_id"], e["name"], c["index"], c["name"],
                c["type"], c["comment"],
            ])

    ws3 = wb.create_sheet("relationships")
    ws3.append(["kind", "from", "from_column", "to", "to_column"])
    for r in relationships:
        ws3.append([
            r["kind"], r["from"], r.get("from_column", ""),
            r["to"], r.get("to_column", ""),
        ])

    # Bold headers and a sensible column width on each sheet
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for col in sheet.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(OUTPUT_XLSX)


def main():
    catalog_file = Path(CATALOG_PATH)
    manifest_file = Path(MANIFEST_PATH)

    if not catalog_file.exists():
        raise SystemExit(f"Could not find {CATALOG_PATH} in {Path.cwd()}")

    catalog = load_json(catalog_file)
    manifest = load_json(manifest_file) if manifest_file.exists() else None
    if manifest is None:
        print(f"(note: {MANIFEST_PATH} not found — relationships will be empty)\n")

    entities = extract_entities(catalog)
    relationships = extract_relationships(manifest)

    # JSON
    Path(OUTPUT_JSON).write_text(
        json.dumps({"entities": entities, "relationships": relationships},
                   indent=2, default=str)
    )
    # CSVs + Excel
    write_csvs(entities, relationships)
    write_excel(entities, relationships)

    print(f"Wrote {OUTPUT_JSON}, {ENTITIES_CSV}, {COLUMNS_CSV}, "
          f"{RELATIONSHIPS_CSV}, {OUTPUT_XLSX}")

    print_summary(entities, relationships)


if __name__ == "__main__":
    main()
